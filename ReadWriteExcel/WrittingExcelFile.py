from openpyxl import load_workbook
wb = load_workbook('Write_Data.xlsx')

sheet = wb.active

#  Write Column Name
Columns = sheet['A1'].value = "Name"
Columns = sheet['B1'].value = "Age"
Columns = sheet['C1'].value = " Salary "
wb.save('Write_Data.xlsx')


print('Successfully written Column Name')


#insert data in columns

name = [ 'Farhina', 'Muskan', 'Diva']
sheet.cell(row=1,column=1).value = " Name "

j = 2
for i in range(0, 3):
    sheet.cell(row=j,column=1).value = name[i]
    #j +=1

    wb.save('Write_Data.xlsx')
print('Successfully written Column data')
