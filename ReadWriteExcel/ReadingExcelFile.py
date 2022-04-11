from openpyxl import load_workbook
wb = load_workbook('Read_Data.xlsx')

sheet = wb.active


# cell header  # excel file row column onojayi read korbe


print(' First cell Name :' ,sheet['A1'].value)
print(' Seceond  cell Name :' ,sheet['B1'].value)

# Arekta system
"""
Cell = sheet.cell(row=1,column=2).value
print(' First cell Name:', Cell)

"""
#Read all data in single code

columns = sheet['A']
for data in columns:
    print(data.value)

    columns = sheet['B']
    for data in columns:
        print(data.value)

